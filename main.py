import engDonustur
import datetime
import os, sys  # Standard Python Libraries
from docxtpl import DocxTemplate, InlineImage  # pip install docxtpl
from docx2pdf import convert
import datetime as tarih
from openpyxl import Workbook, load_workbook
import mailGonder

wb = load_workbook("liste.xlsx")
ws = wb.active

# Aktif çalışma sayfasının adını yazdırma
# print(wb.sheetnames)
for veri in range(2,5):  # 128 kişi olacak
    adiSoyadi = ws.cell(veri, 2).value
    kisiAdi = adiSoyadi
    universitesi = ws.cell(veri, 3).value
    gonderilecekMail = ws.cell(veri, 4).value
    baslik = ws.cell(veri, 5).value
    os.chdir(sys.path[0])
    doc = DocxTemplate("Template.docx")
    date = tarih.datetime.now()
    context = {
        "adsoyad": adiSoyadi,
        "baslik": baslik,
        "universite": universitesi,
        "tarih": "28 February 2022",
    }

    doc.render(context)
    adiSoyadi = engDonustur.engDonustur(adiSoyadi)
    dokumanAdi = str(veri - 1) + "_" + adiSoyadi + " IIC2022 " + "Kabul Mektubu" + ".docx";
    doc.save("word/" + dokumanAdi)

    dosyaadi = str(veri - 1) + "_" + adiSoyadi + " IIC2022 " + "Kabul Mektubu" + ".pdf"
    convert("word/" + dokumanAdi, "pdf/" + dosyaadi)
    print(kisiAdi + " Adlı Kişiye Eposta Gönderilecektir.")
    #mailGonder.mailGonder(gonderilecekMail, dosyaadi)

print("Tüm İşlem Tamamlandı.")
